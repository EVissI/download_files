from datetime import datetime
from typing import Optional
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from loguru import logger
from sqlalchemy.exc import SQLAlchemyError

from bot.common.func.func import determine_rank
from bot.db.dao import UserDAO, DetailedAnalysisDAO

from typing import TYPE_CHECKING
from fluentogram import TranslatorRunner
from bot.common.utils.i18n import get_all_locales_for_key

if TYPE_CHECKING:
    from locales.stub import TranslatorRunner


async def generate_detailed_analysis_report(
    dao: DetailedAnalysisDAO,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
) -> io.BytesIO:
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Детальный анализ"

        # Стили
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )
        center_align = Alignment(horizontal="center")

        # Заголовки
        headers = [
            "ID анализа",
            "ID пользователя",
            "Имя игрока",
            "Дата анализа",
            "Плохие ходы",
            "Очень плохие ходы",
            "Ошибка (Chequerplay)",
            "Рейтинг Chequerplay",
            "Очень удачные броски",
            "Удачные броски",
            "Неудачные броски",
            "Очень неудачные броски",
            "Рейтинг удачи",
            "Рейтинг кубика",
            "Пропущенные даблы ниже CP",
            "Пропущенные даблы выше CP",
            "Неправильные даблы ниже SP",
            "Неправильные даблы выше TG",
            "Ошибки взятия",
            "Неправильные пасы",
            "Ошибки куба",
            "Общая ошибка",
            "Общий рейтинг",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        analyses = await dao.get_all_detailed_analyzes(start_date, end_date)
        current_row = 2

        for analysis in analyses:
            ws.cell(row=current_row, column=1).value = analysis.id
            ws.cell(row=current_row, column=2).value = analysis.user_id
            ws.cell(row=current_row, column=3).value = analysis.player_name
            ws.cell(row=current_row, column=4).value = (
                analysis.created_at.strftime("%Y-%m-%d")
                if analysis.created_at
                else "N/A"
            )
            ws.cell(row=current_row, column=5).value = analysis.moves_marked_bad
            ws.cell(row=current_row, column=6).value = analysis.moves_marked_very_bad
            ws.cell(row=current_row, column=7).value = analysis.error_rate_chequer
            ws.cell(row=current_row, column=8).value = analysis.chequerplay_rating
            ws.cell(row=current_row, column=9).value = analysis.rolls_marked_very_lucky
            ws.cell(row=current_row, column=10).value = analysis.rolls_marked_lucky
            ws.cell(row=current_row, column=11).value = analysis.rolls_marked_unlucky
            ws.cell(row=current_row, column=12).value = (
                analysis.rolls_marked_very_unlucky
            )
            ws.cell(row=current_row, column=13).value = analysis.luck_rating
            ws.cell(row=current_row, column=14).value = analysis.cube_decision_rating
            ws.cell(row=current_row, column=15).value = analysis.missed_doubles_below_cp
            ws.cell(row=current_row, column=16).value = analysis.missed_doubles_above_cp
            ws.cell(row=current_row, column=17).value = analysis.wrong_doubles_below_sp
            ws.cell(row=current_row, column=18).value = analysis.wrong_doubles_above_tg
            ws.cell(row=current_row, column=19).value = analysis.wrong_takes
            ws.cell(row=current_row, column=20).value = analysis.wrong_passes
            ws.cell(row=current_row, column=21).value = analysis.cube_error_rate
            ws.cell(row=current_row, column=22).value = analysis.snowie_error_rate
            ws.cell(row=current_row, column=23).value = analysis.overall_rating
            current_row += 1

        column_widths = {
            "A": 25,
            "B": 25,
            "C": 25,
            "D": 25,
            "E": 25,
            "F": 25,
            "G": 25,
            "H": 25,
            "I": 25,
            "J": 25,
            "K": 25,
            "L": 25,
            "M": 35,
            "N": 35,
            "O": 35,
            "P": 35,
            "Q": 35,
            "R": 25,
            "S": 25,
            "T": 25,
            "U": 25,
            "V": 25,
            "W": 25,
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        logger.info(f"Сгенерирован Excel отчет с {current_row-2} записями")
        return excel_buffer

    except Exception as e:
        logger.error(f"Ошибка при генерации Excel отчета: {e}")
        raise


async def generate_detailed_user_analysis_report(
    dao: DetailedAnalysisDAO,
    player_name: str = None,
    start_date: datetime = None,
    end_date: datetime = None,
) -> io.BytesIO:
    """
    Генерирует Excel отчет со статистикой детального анализа по player_name.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"Детальный анализ {player_name or ''}"

    # Стили
    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
    )
    center_align = Alignment(horizontal="center")

    # Заголовки
    headers = [
        "ID анализа",
        "ID пользователя",
        "Имя игрока",
        "Дата анализа",
        "Плохие ходы",
        "Очень плохие ходы",
        "Ошибка (Chequerplay)",
        "Рейтинг Chequerplay",
        "Очень удачные броски",
        "Удачные броски",
        "Неудачные броски",
        "Очень неудачные броски",
        "Рейтинг удачи",
        "Рейтинг кубика",
        "Пропущенные даблы ниже CP",
        "Пропущенные даблы выше CP",
        "Неправильные даблы ниже SP",
        "Неправильные даблы выше TG",
        "Ошибки взятия",
        "Неправильные пасы",
        "Ошибки куба",
        "Общая ошибка",
        "Общий рейтинг",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    try:
        if not player_name:
            raise ValueError("Не указано игровое имя (player_name)")

        analyses = await dao.get_detailed_analyzes_by_player_name(
            player_name, start_date, end_date
        )
        current_row = 2

        for analysis in analyses:
            ws.cell(row=current_row, column=1).value = analysis.id
            ws.cell(row=current_row, column=2).value = analysis.user_id
            ws.cell(row=current_row, column=3).value = analysis.player_name
            ws.cell(row=current_row, column=4).value = (
                analysis.created_at.strftime("%Y-%m-%d")
                if analysis.created_at
                else "N/A"
            )
            ws.cell(row=current_row, column=5).value = analysis.moves_marked_bad
            ws.cell(row=current_row, column=6).value = analysis.moves_marked_very_bad
            ws.cell(row=current_row, column=7).value = analysis.error_rate_chequer
            ws.cell(row=current_row, column=8).value = analysis.chequerplay_rating
            ws.cell(row=current_row, column=9).value = analysis.rolls_marked_very_lucky
            ws.cell(row=current_row, column=10).value = analysis.rolls_marked_lucky
            ws.cell(row=current_row, column=11).value = analysis.rolls_marked_unlucky
            ws.cell(row=current_row, column=12).value = (
                analysis.rolls_marked_very_unlucky
            )
            ws.cell(row=current_row, column=13).value = analysis.luck_rating
            ws.cell(row=current_row, column=14).value = analysis.cube_decision_rating
            ws.cell(row=current_row, column=15).value = analysis.missed_doubles_below_cp
            ws.cell(row=current_row, column=16).value = analysis.missed_doubles_above_cp
            ws.cell(row=current_row, column=17).value = analysis.wrong_doubles_below_sp
            ws.cell(row=current_row, column=18).value = analysis.wrong_doubles_above_tg
            ws.cell(row=current_row, column=19).value = analysis.wrong_takes
            ws.cell(row=current_row, column=20).value = analysis.wrong_passes
            ws.cell(row=current_row, column=21).value = analysis.cube_error_rate
            ws.cell(row=current_row, column=22).value = analysis.snowie_error_rate
            ws.cell(row=current_row, column=23).value = analysis.overall_rating
            current_row += 1

        column_widths = {
            "A": 25,
            "B": 25,
            "C": 25,
            "D": 25,
            "E": 25,
            "F": 25,
            "G": 25,
            "H": 25,
            "I": 25,
            "J": 25,
            "K": 25,
            "L": 25,
            "M": 35,
            "N": 35,
            "O": 35,
            "P": 35,
            "Q": 35,
            "R": 25,
            "S": 25,
            "T": 25,
            "U": 25,
            "V": 25,
            "W": 25,
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        logger.info(
            f"Сгенерирован Excel отчет с {current_row-2} записями для {player_name}"
        )
        return excel_buffer

    except Exception as e:
        logger.error(f"Ошибка при генерации Excel отчета: {e}")
        raise
